from openpyxl import load_workbook


def dataGetter(sheetname):
    wb=load_workbook(r"C:\Users\AZ\POM2\excel\info.xlsx")
    sheet=wb[sheetname]
    tot_row=sheet.max_row
    tot_col=sheet.max_column
    main=[]
    for i in range(2,tot_row+1):
        temp=[]
        for j in range(1,tot_col+1):
            data=sheet.cell(row=i,column=j).value
            temp.insert(j,data)
        main.insert(i,temp)
    return main